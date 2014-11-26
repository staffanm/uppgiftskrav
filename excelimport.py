#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import print_function, unicode_literals
from collections import OrderedDict, defaultdict
import json
import re
import sys

from openpyxl import load_workbook

from register.utils import parse_lagrum, format_lagrum, format_link

assert sys.version_info < (
    3, 0, 0), "String handling not Py3 compliant, plz rewrite"


def myndighet_id(d, key):
    assert key
    if key not in d:
        d[key] = {'id': len(d) + 1,
                  'name': key}
    return d[key]['id']


def value(row, idx, max_length=None):
    val = row[idx].internal_value
    if isinstance(val, float):
        return unicode(int(val))  # 0.0 -> '0'
    elif not val:
        return ""
    if not isinstance(val, unicode):
        val = unicode(val)
    if max_length:
        val = val[:max_length]
    return val.strip()


fsrefs = defaultdict(int)


def lagrum(row, idx):
    val = value(row, idx)
    ds = parse_lagrum(val)
    for d in ds:
        if 'fs' in d:
            fsrefs[(d['fs'], d['fsnr'])] += 1
    res = format_lagrum(ds)
    # print("%s => %s => %s" % (val, ds, res))
    return res

# make note of in a {value: id} dict
def note(row, idx, dictionary):
    val = value(row, idx)
    if val not in dictionary and val:
        dictionary[val] = len(dictionary) + 1
    return val


def note2(row, idx, idx2, dictionary):
    val = value(row, idx)
    val2 = value(row, idx2)
    if (val, val2) not in dictionary and val:
        dictionary[(val, val2)] = len(dictionary) + 1
    return val


def choice(row, idx, choices, default=None, warn=True):
    val = value(row, idx)
    lval = val.lower()
    lchoices = [x.lower() if hasattr(x, 'lower') else x for x in choices]
    if lval in lchoices:
        return lchoices.index(lval)
    else:
        if val and warn:
            print("WARNING (%s): %r not in %r" %
                  (row[3].internal_value, val, choices))
        if default is not None:  # must be able to be 0 or other falsieness
            return default
        else:
            raise KeyError(val)


# just wrap the result of choice in a list for the time being
def manychoice(row, idx, choices):
    try:
        return [choice(row, idx, choices, warn=False)]
    except KeyError:
        return []


def integer(row, idx):
    val = value(row, idx)
    if val and val.isdigit():
        return int(val)
    # else return None

def has(row, idx, sought_value):
    val = value(row, idx)
    return "x" in val.lower()

def nbf(row, idx):
    """Given a row and index to that row, return 0 if it contains "Nej", 1
    if it contains "Ja", and -1 otherwise (eg empty).

    """
    
    return choice(row, idx, ('Nej', 'Ja'), -1)
    # val = value(row, idx)
    # if val:
    #    return val == "Ja"
    # else:
    #    return None


def many(row, idx, filter=None):
    """Given a row and a column index to that row, extract the value from
    that cell and split the contents into a comma separated list.

    :param row: A excel row
    :param idx: A numerical index representing the ordinal of the sought column
    :param filter: A callable that returns True or False for each element in the extracted list. Only elements returning True will be returned. If None, accepts all values.
    """
    val = value(row, idx)
    if filter is None:
        filter = lambda x: True
    if val:
        l = val.split(",")
        return [[x.strip().upper()] for x in l if x.strip() and filter(x)]
    else:
        return []

def avgransad(k):
    if k['etjanst'] == 1:
        return False
    if k['volymer_2012'] < 10 and k['antal_foretag'] < 10:
        return True
    if k['svarighet_etjanst'] == 6:  # endast maskin-till-maskin
        return True
    if k['svarighet_ej_etjanst'] == 1 and k['etjanst'] != 1: # ingen info på myndhemsida, ingen etjänst
        return True
    return False




def make_fixture(newdata, olddata, fixture):
    # 1: Extract all Kravs (and create a list of Myndigheter as
    # django.contrib.auth.models.Group as well)
    print("Opening new data %s" % newdata)
    ws = load_workbook(newdata, use_iterators=True).get_active_sheet()

    # these could be set dynamically by examining ws.rows[0]
    VERKSAMHETSOMRADE = 0
    ANSVARIG_MYNDIGHET = 1
    KARTLAGGANDE_MYNDIGHET = 2
    ID = 3
    NAMN = 4  # UPPGIFTSKRAV
    FORFATTNING = 5  # deprecated
    PARAGRAF = 6  # deprecated
    LAGRUM = 7
    URSPRUNG = 8
    BESKRIVNING = 9  # deprecated
    ANTECKNING = 10
    KORT_BESKRIVNING = 11
    LEDER_TILL_INSAMLING = 12
    EGNA_NOTERINGAR = 13
    KALENDERSTYRT = 14
    PERIODICITET = 15
    HANDELSESTYRT = 16
    INITIERANDE_PART = 17
    OVRIGT_NAR = 18
    BRANSCH = 19
    ARBETSGIVARE = 20
    FORETAGSFORM = 21
    STORLEK = 22
    STORLEKSKRITERIER = 23
    OVRIGA_URVALSKRITERIER = 24
    ANTAL_FORETAG = 25
    ANNAN_INGIVARE = 26
    UNDERSKRIFT = 27
    ETJANST = 28
    SVARIGHET_EJ_ETJANST = 29
    SVARIGHET_ETJANST = 30
    VOLYMER_TIDIGARE = 31
    VOLYMER_2012 = 32
    VOLYMER_ETJANST = 33
    OVRIGT_HUR = 34

    krav = OrderedDict()
    uppgift = OrderedDict()
    verksamhetsomrade = OrderedDict()
    myndighet = OrderedDict()
    myndighet['Företagsdatanämnden'] = 1

    print("Enumerating Krav")
    avgcnt = 0
    for (cnt, row) in enumerate(ws.iter_rows()):
        if cnt < 1 or not row[ID].internal_value:
            continue
        k = {
            'id': cnt,
            'kravid': value(row, ID, max_length=7),
            'ansvarig_myndighet': [note(row, ANSVARIG_MYNDIGHET, myndighet)],
            'kartlaggande_myndighet': [note(row, KARTLAGGANDE_MYNDIGHET, myndighet)],
            'verksamhetsomrade': [note2(row, VERKSAMHETSOMRADE, KARTLAGGANDE_MYNDIGHET, verksamhetsomrade)],
            'namn': value(row, NAMN),
            'forfattning': value(row, FORFATTNING, max_length=50),
            'paragraf': value(row, PARAGRAF, max_length=50),
            'lagrum': value(row, LAGRUM),
            'ursprung': choice(row, URSPRUNG, (False, "Nationellt", "EU mm"), default=-1),
            'beskrivning': value(row, BESKRIVNING),
            'anteckning': value(row, ANTECKNING),
            'kortbeskrivning': value(row, KORT_BESKRIVNING),
            'leder_till_insamling': nbf(row, LEDER_TILL_INSAMLING),
            'egna_noteringar': value(row, EGNA_NOTERINGAR),
            'kalenderstyrt': nbf(row, KALENDERSTYRT),
            'periodicitet': manychoice(row, PERIODICITET, (False, 'Januari', 'Februari', 'Mars', 'April', 'Maj', 'Juni', 'Juli', 'Augusti', 'September', 'Oktober', 'November', 'December', 'Veckovis', 'Månadsvis', 'Kvartalsvis', 'Årsvis (ej särskilt datum)')),
            'handelsestyrt': nbf(row, HANDELSESTYRT),
            'initierande_part': choice(row, INITIERANDE_PART, (False, 'Myndigheten', 'Företaget'), default=-1, warn=False),
            'ovrigt_nar': value(row, OVRIGT_NAR),
            'beror_bransch': not(has(row, BRANSCH, "x")), # False if contains "X", True otherwise
            'bransch': many(row, BRANSCH, lambda val: val.lower() != "x"), # filter out "X"
            'arbetsgivare': nbf(row, ARBETSGIVARE),
            'beror_foretagsform': not(has(row, FORETAGSFORM, "x")), # False if contains "X", True otherwise
            'foretagsform': many(row, FORETAGSFORM, lambda val: val.lower() != "x"),# filter out "X"
            'storlek': nbf(row, STORLEK),
            'storlekskriterier': value(row, STORLEKSKRITERIER),
            'ovriga_urvalskriterier': value(row, OVRIGA_URVALSKRITERIER),
            'antal_foretag': integer(row, ANTAL_FORETAG),
            'annan_ingivare': nbf(row, ANNAN_INGIVARE),
            'underskrift': nbf(row, UNDERSKRIFT),
            'etjanst': nbf(row, ETJANST),
            'svarighet_ej_etjanst': choice(row, SVARIGHET_EJ_ETJANST, ('0', '1', '2', '3', '4'), default=-1),
            'svarighet_etjanst': choice(row, SVARIGHET_ETJANST, (False, False, False, False, False, '5', '6', '7'), default=-1),
            'volymer_tidigare': integer(row, VOLYMER_TIDIGARE),
            'volymer_2012': integer(row, VOLYMER_2012),
            'volymer_etjanst': integer(row, VOLYMER_ETJANST),
            'ovrigt_hur': value(row, OVRIGT_HUR),
            'uppgifter': []
        }
        if (not k['leder_till_insamling']):
            # not a Uppgiftskrav, don't import this at all
            sys.stdout.write("X")
        else:
            k['avgransad'] = avgransad(k)
            if k['avgransad']:
                avgcnt += 1 
            sys.stdout.write("A" if k['avgransad'] else ".")
            krav[row[ID].internal_value] = k
            
        sys.stdout.flush()
    print("\nDone, %s krav (%s avgransade, %s myndigheter) in %s" %
          (len(krav), avgcnt, len(myndighet), newdata))

    # 2: Find information about Grundläggande uppgifter (GU)
    print("Opening old data %s" % olddata)
    ws = load_workbook(olddata, use_iterators=True).get_active_sheet()
    print("Enumerating Uppgifter")
    rows = ws.iter_rows()
    next(rows)  # 1st row contains no usable data
    firstrow = next(rows)
    for (cnt, cell) in enumerate(firstrow[30:]):
        uppgift[cnt] = {'id': cnt + 1,
                        'uppgiftid': 'UD%04d' % (cnt + 1),
                        'namn': cell.internal_value}

    print("Done, found %s uppgifter" % len(uppgift))

    # 3: connect each Krav to a number of GU
    print("Connecting uppgifter to krav")
    for row in rows:
        try:
            kravid = row[ID].internal_value
            d = krav[kravid]
        except KeyError:
            # quite a common problem, olddata contains 2409 rows, newdata contains 1422 rows
            # print("Old data contained krav %s which is not present in new data" % row[ID].internal_value)
            continue

        d['uppgifter'] = [
            cnt + 1 for (cnt, x) in enumerate(row[30:]) if x.internal_value in ('Obligatorisk', 'Frivillig')]
        sys.stdout.write(".")
        sys.stdout.flush()

    print("\nDone, now reformatting to fixture format")
    # 4: reformat everything to appropriate json dicts
 
    # 4.1 hardcode Bransch / företagsform / periodicitet
    data = [{'model': 'register.bransch',
             'pk': 1,
             'fields': {'snikod': 'A',
                        'beskrivning': 'Jordbruk, skogsbruk och fiske'}},
            {'model': 'register.bransch',
             'pk': 2,
             'fields': {'snikod': 'B',
                        'beskrivning': 'Utvinning av mineral'}},
            {'model': 'register.bransch',
             'pk': 3,
             'fields': {'snikod': 'C',
                        'beskrivning': 'Tillverkning'}},
            {'model': 'register.bransch',
             'pk': 4,
             'fields': {'snikod': 'D',
                        'beskrivning': 'Försörjning av el, gas, värme och kyla'}},
            {'model': 'register.bransch',
             'pk': 5,
             'fields': {'snikod': 'E',
                        'beskrivning': 'Vattenförsörjning; avloppsrening, avfallshantering och sanering'}},
            {'model': 'register.bransch',
             'pk': 6,
             'fields': {'snikod': 'F',
                        'beskrivning': 'Byggverksamhet'}},
            {'model': 'register.bransch',
             'pk': 7,
             'fields': {'snikod': 'G',
                        'beskrivning': 'Handel; reparation av motorfordon och motorcyklar'}},
            {'model': 'register.bransch',
             'pk': 8,
             'fields': {'snikod': 'H',
                        'beskrivning': 'Transport och magasinering'}},
            {'model': 'register.bransch',
             'pk': 9,
             'fields': {'snikod': 'I',
                        'beskrivning': 'Hotell- och restaurangverksamhet'}},
            {'model': 'register.bransch',
             'pk': 10,
             'fields': {'snikod': 'J',
                        'beskrivning': 'Informations- och kommunikationsverksamhet'}},
            {'model': 'register.bransch',
             'pk': 11,
             'fields': {'snikod': 'K',
                        'beskrivning': 'Finans- och försäkringsverksamhet'}},
            {'model': 'register.bransch',
             'pk': 12,
             'fields': {'snikod': 'L',
                        'beskrivning': 'Fastighetsverksamhet'}},
            {'model': 'register.bransch',
             'pk': 13,
             'fields': {'snikod': 'M',
                        'beskrivning': 'Verksamhet inom juridik, ekonomi, vetenskap och teknik'}},
            {'model': 'register.bransch',
             'pk': 14,
             'fields': {'snikod': 'N',
                        'beskrivning': 'Uthyrning, fastighetsservice, resetjänster och andra stödtjänster'}},
            {'model': 'register.bransch',
             'pk': 15,
             'fields': {'snikod': 'O',
                        'beskrivning': 'Offentlig förvaltning och försvar; obligatorisk socialförsäkring'}},
            {'model': 'register.bransch',
             'pk': 16,
             'fields': {'snikod': 'P',
                        'beskrivning': 'Utbildning'}},
            {'model': 'register.bransch',
             'pk': 17,
             'fields': {'snikod': 'Q',
                        'beskrivning': 'Vård och omsorg; sociala tjänster'}},
            {'model': 'register.bransch',
             'pk': 18,
             'fields': {'snikod': 'R',
                        'beskrivning': 'Kultur, nöje och fritid'}},
            {'model': 'register.bransch',
             'pk': 19,
             'fields': {'snikod': 'S',
                        'beskrivning': 'Annan serviceverksamhet'}},
            {'model': 'register.bransch',
             'pk': 20,
             'fields': {'snikod': 'T',
                        'beskrivning': 'Förvärvsarbete i hushåll; hushållens produktion av diverse varor och tjänster för eget bruk'}},
            {'model': 'register.bransch',
             'pk': 21,
             'fields': {'snikod': 'U',
                        'beskrivning': 'Verksamhet vid internationella organisationer, utländska ambassader o.d.'}},
            {'model': 'register.foretagsform',
             'pk': 1,
             'fields': {'formkod': 'E',
                        'beskrivning': 'Enskild näringsidkare'}},
            {'model': 'register.foretagsform',
             'pk': 2,
             'fields': {'formkod': 'AB',
                        'beskrivning': 'Aktiebolag'}},
            {'model': 'register.foretagsform',
             'pk': 3,
             'fields': {'formkod': 'HB',
                        'beskrivning': 'Handelsbolag'}},
            {'model': 'register.foretagsform',
             'pk': 4,
             'fields': {'formkod': 'KB',
                        'beskrivning': 'Kommanditbolag'}},
            {'model': 'register.foretagsform',
             'pk': 5,
             'fields': {'formkod': 'BRF',
                        'beskrivning': 'Bostadsrättsförening'}},
            {'model': 'register.foretagsform',
             'pk': 6,
             'fields': {'formkod': 'EK',
                        'beskrivning': 'Ekonomisk förening'}},
            {'model': 'register.foretagsform',
             'pk': 7,
             'fields': {'formkod': 'A',
                        'beskrivning': 'Annan företagsform'}},
            {'model': 'register.periodicitet',
             'pk': 1,
             'fields': {'beskrivning': 'Januari'}},
            {'model': 'register.periodicitet',
             'pk': 2,
             'fields': {'beskrivning': 'Februari'}},
            {'model': 'register.periodicitet',
             'pk': 3,
             'fields': {'beskrivning': 'Mars'}},
            {'model': 'register.periodicitet',
             'pk': 4,
             'fields': {'beskrivning': 'April'}},
            {'model': 'register.periodicitet',
             'pk': 5,
             'fields': {'beskrivning': 'Maj'}},
            {'model': 'register.periodicitet',
             'pk': 6,
             'fields': {'beskrivning': 'Juni'}},
            {'model': 'register.periodicitet',
             'pk': 7,
             'fields': {'beskrivning': 'Juli'}},
            {'model': 'register.periodicitet',
             'pk': 8,
             'fields': {'beskrivning': 'Augusti'}},
            {'model': 'register.periodicitet',
             'pk': 9,
             'fields': {'beskrivning': 'September'}},
            {'model': 'register.periodicitet',
             'pk': 10,
             'fields': {'beskrivning': 'Oktober'}},
            {'model': 'register.periodicitet',
             'pk': 11,
             'fields': {'beskrivning': 'November'}},
            {'model': 'register.periodicitet',
             'pk': 12,
             'fields': {'beskrivning': 'December'}},
            {'model': 'register.periodicitet',
             'pk': 13,
             'fields': {'beskrivning': 'Veckovis'}},
            {'model': 'register.periodicitet',
             'pk': 14,
             'fields': {'beskrivning': 'Månadsvis'}},
            {'model': 'register.periodicitet',
             'pk': 15,
             'fields': {'beskrivning': 'Kvartalsvis'}},
            {'model': 'register.periodicitet',
             'pk': 16,
             'fields': {'beskrivning': 'Årsvis (ej särskilt datum)'}},
            ]

    for (name, myndname), id in verksamhetsomrade.items():
        data.append({'model': 'register.verksamhetsomrade',
                     'pk': id,
                     'fields': {'omrade': name,
                                'myndighet' :  [myndname]}})

    for u in uppgift.values():
        data.append({'model': 'register.uppgift',
                     'pk': u['id'],
                     'fields': {'uppgiftid': u['uppgiftid'],
                                'namn': u['namn']}
                     })

    for k in krav.values():
        ck = k.copy()
        del ck['id']
        # ck['myndighet']
        # ck['uppgifter']
        # ck['url'] = '' # not currently extracted from excel data
        data.append({'model': 'register.krav',
                     'pk': k['id'],
                     'fields': ck
                     })

    # 5: aaaand we're done
    print("Serializing to JSON as %s" % fixture)
    with open(fixture, "w") as fp:
        json.dump(data, fp, indent=4)
    print("Wrote %s entries to %s" % (len(data), fixture))

    # the below code attempted to map FS refs to lawyer-readable labels
    #     res = "{"
    #     with open("sfstitles.nt") as fp:
    #         for line in fp:
    #             fsnr, title = re.match(
    #                 '<http://rinfo.lagrummet.se/publ/sfs/(\d+:[^>]+)> <http://purl.org/dc/terms/title> "([^"]*)"', line).groups()
    #         res += '"%s": "%s"' % (fsnr, title)
    #     res += "}"
    #     sfstitles = json.loads(res)
    # 
    #     for k, v in fsrefs.items():
    #         fs, fsnr = k
    #         if fs == "SFS":
    #             name = sfstitles[fsnr]
    #         else:
    #             name = "Myndighetens föreskrifter (%s) om X" % fsnr
    #     print("(%r, %r): '%s'" % (k, v, name))


if __name__ == "__main__":

    if len(sys.argv) != 4:
        print(
            "Usage: %s [new excel data] [old excel data] [fixture]" % sys.argv[0])
        sys.exit()

    newdata = sys.argv[1]
    olddata = sys.argv[2]
    fixture = sys.argv[3]
    make_fixture(newdata, olddata, fixture)
